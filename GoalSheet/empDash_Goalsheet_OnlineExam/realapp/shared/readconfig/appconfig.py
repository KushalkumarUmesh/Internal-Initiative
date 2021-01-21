import openpyxl

class AppConfig(object) :
    def __init__(self, configInfile) :
        self.configInfile = configInfile
        self.attribute = {}  # Dict of Attributes
        self.readConfigFile() 
        
    #Reading the file, and save keyValue pairs, convert everything into String
    def readConfigFile(self) :
        conf_wb = openpyxl.load_workbook(self.configInfile)
        ws = conf_wb.worksheets[0]
        c_row = 0
        circuitBreaker = 100
        while circuitBreaker :
            c_row += 1
            circuitBreaker -= 1
            c = ws.cell(row=c_row, column=1).value
            v = ws.cell(row=c_row, column=2).value # Value
            if c :
                c=str(c) # Convert to String, just to be sure
                c= c.strip() # Remove white space
                v=str(v)  # All Values are also treated as strings, its for the program to ensure that its correct number
                v= v.strip() # Remove white space
                if c == "END" : break
                if v :
                    self.attribute[c] = v
                else :
                    print("Variable{0} is does not have a value.format(c)")
                    return(1)
            else :
                continue
        return (0)