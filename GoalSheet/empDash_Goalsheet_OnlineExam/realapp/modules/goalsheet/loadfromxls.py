"""
K.Srinivas, 9-Jun-2018

Project: Goal Sheet
Description: One time use application to read and update "target-set" field in the DB. This file is ONLY
for reading and processing the XLS-files. The DB-part is handled in loadtargetview.py
"""


import openpyxl
import glob
import os
import re
import shutil
import pathlib
cwd=os.getcwd()
basedir = "C:\\goalsheetsfromvandana"
processeddir = "\\processed"
dcdir = "C:\\goalsheetsfromvandana\\DC Lead_ GOAL SHEET_2018-19"
nondcdir = "C:\\goalsheetsfromvandana\\GOAL SHEET_2018-19"
filetemp="*.xlsx"


#TODO:
#fill correct goalTargetCell and goalDCTargetCell : Hard-coded Cell-numbers
cellEmpNo = "E7" # List out Cell numbers
#DC-Goal-Sheet
import openpyxl
goalDCTargetCell = {3:"F27", 4:"F31",5:"F32",6:"F33",7:"F34",8:"F35", 9:"F37",13:"F38",12:"F39", 
                   14:"F41",15:"F42",16:"F43",17:"F44",18:"F45",
                   19:"F55",20:"F56",21:"F57",22:"F58",23:"F59",24:"F60",25:"F61"
                   } # 
#Emp-Goal-Sheet
goalTargetCell = {1:"F27", 4:"F31",5:"F32",6:"F33",7:"F34",8:"F35", 9:"F37",13:"F38",12:"F39", 
                   14:"F41",15:"F42",16:"F43",17:"F44",18:"F45"
                   } # 

#Method to get row and col numbers for accessing the call based on String e.g. E10 = 6,10
def getCellRowCol(cellNo) :
    c_row=int(cellNo[1:])
    c_col=int(cellNo[0],36)-9
    return (c_row, c_col)


def readGoalSheetFile(filename,dc=False) :
    targetsSet = dict() # Key is the goal-number, value is the text in the cell
    bumpCol = 0
    conf_wb = openpyxl.load_workbook(filename)
    ws = conf_wb.worksheets[0]

    #Get Employee Number
    (c_row,c_col)  = getCellRowCol(cellEmpNo)
    empNo = ws.cell(row=c_row, column=c_col).value
    if not isinstance(empNo, int) : # Eiter not present OR is not a number, try next Col
        empNo = ws.cell(row=c_row, column=c_col+1).value
        if not isinstance(empNo, int) : # Eiter not present OR is not a number, try next Col
            print("Emp Not found in file Processing File:%s" % ( filename))
            return ("", None)
        else : # Number was found in F column instead of E
            c_col += 1 # Bump it up
            bumpCol = 1

    empName = str(ws.cell(row=c_row, column=c_col-1).value) + " " +  str(ws.cell(row=c_row, column=c_col-2).value)
    #Add Checks, assignment if needed
#    print("Processing Emp:%s : %s" % (empNo, empName))
    useDict = goalTargetCell
    if dc : useDict = goalDCTargetCell

    for k in useDict.keys() :
        (c_row,c_col)  = getCellRowCol(useDict[k])
        c_col += bumpCol # In case someone used the older template
        cellVal = ws.cell(row=c_row, column=c_col).value
        targetsSet[k] = cellVal
#    print("Targets are:" + str(targetsSet))
    return (targetsSet, empNo)

def readGoalFiles(dc=False) :
    if dc :
        flist = glob.glob(dcdir + '/*.xlsx', recursive=True) 
    else :
        flist = glob.glob(nondcdir + '/*/*.xlsx', recursive=True)     
    return flist

def moveFileToProcessed(fname, dc) :
    baseName = os.path.basename(fname)
    newDir = basedir + processeddir + fname[len(basedir):-len(baseName)-1 ]
    newName = newDir + "\\" + baseName
#    print("Making Dir:" + newDir)
    pathlib.Path(newDir).mkdir(parents=True, exist_ok=True) 
#    print("Moving:" + fname + " to " + newName)
    os.rename(fname,newName)
    return
    

if __name__ == "__main__":
    #Read each XLS
    #Get the emp.no
    #Check if emp.no is present
        #Check if emp is HRMS-db
            #Else Log - file-name, emp.no
            #Move file to another folder?
        #Else assign goal-sheet. Is it possible based on file-name?
    #assign the targets
    #List out those who are not present in DB
    ### ERROR CHECKING
    # List employees in DB who don't have a goal-sheet
    # List DC_leads-by goal-sheet??? Is it possible
    pass



"""
class AppConfig(object) :
    def __init__(self, configInfile) :
        self.configInfile = configInfile
        self.attribute = {}  # Dict of Attributes
        self.readConfigFile() 
        
    #Reading the file, and save keyValue pairs, convert everything into String
    def readConfigFile(self) :
        conf_wb = openpyxl.load_workbook(self.configInfile)
        ws = conf_wb.worksheets[0]
        c_row = 0
        circuitBreaker = 100
        while circuitBreaker :
            c_row += 1
            circuitBreaker -= 1
            c = ws.cell(row=c_row, column=1).value
            v = ws.cell(row=c_row, column=2).value # Value
            if c :
                c=str(c) # Convert to String, just to be sure
                c= c.strip() # Remove white space
                v=str(v)  # All Values are also treated as strings, its for the program to ensure that its correct number
                v= v.strip() # Remove white space
                if c == "END" : break
                if v :
                    self.attribute[c] = v
                else :
                    print("Variable{0} is does not have a value.format(c)")
                    return(1)
            else :
                continue
        return (0)
"""