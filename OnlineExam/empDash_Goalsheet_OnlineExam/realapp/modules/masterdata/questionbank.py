"""
Project: OnlineExam
Domain: QuestionBank
Description: The QuestionBank Domain handles all functionalities related to questions, answers, topics, subtopics, etc.
The following need to be supplied. Responses are expected to be JSON or dict, depending on a global flag. JSON is for Angular2
UI. Ability to return the data as-is allows for alternate processing e.g. storing in DB
a)-List of Topics (Questions) - This is collected from all the question-banks, check needs to made for Unique
b)-List of SubTopics (Questions) - This is collected from all the question-banks, check needs to made for Unique
h)-List of Question-Banks (Banks with Description) (Questions) 
i)-List of Questions (without Answers, exam) (Questions) : Randomize, limit to a number, select level
j)-List of Questions (with Answers and Description) (Questions)

Author: K.Srinivas
16-Mar-2018
"""


# In[1]:
import logging

import glob
import openpyxl
import pandas as pd
import sys
import json
sys.path.insert(0,'..')
sys.path.insert(0,'../../shared/readconfig')

from appconfig import AppConfig

# This class reads reads all the question bank-XLS files and collates data for future use
# Data is expected to be small enough to be read and stored in memory. In case the memory is too much,
# Only data that is needed is the header-data, actual questions can be re-read on-demand
# I use appConfig heavily. All the paths, values needed must be defined in Appconfig
class QuestionBank(object) :
    def __init__(self, appConf) :
        qmatch = appConf.attribute["QuestionMatch"]  # directory containing all question-bank files
        tRow = 2 # appConf.attribute["TOPICROW"] # Row to find TOPIC in the XLS-file, will assume the remaining two
        tCol = 2 # appConf.attribute["TOPICROW"] 
        self.qBanksAll = dict() # All question banks stored
        self.topics = [] # List of all topics
        self.subtopicsList = dict() # Key is TOPIC, Values is a list of subtopics
        self.subtopics = dict() # List of all subtopics Key= topic, value=[subtopic, description, df]
        self.subtopicsDF = dict() # List of all subtopics Key= topic, value=[subtopic, description, df]
        
        for file in glob.glob(qmatch):  # For each of the question banks
            qbank = dict() # Define a question Bank
            #Get title, Get Subtitle, Get Desc
            conf_wb = openpyxl.load_workbook(file)
            ws = conf_wb.worksheets[0] # Assume only 1 worksheet
            if len(conf_wb.worksheets) > 1 : # Error Check
                print("More than one worksheets found in %s" % (file))
                
            c_row =  tRow
            qbank["TOPIC"] = ws.cell(row=c_row, column=tCol).value
            qbank["SUBTOPIC"] = ws.cell(row=(c_row+1), column=tCol).value # Value
            qbank["DESCRIPTION"] = ws.cell(row=(c_row+2), column=tCol).value # Value
            conf_wb.close() # Close it here, just to be sure
            
            df = pd.read_excel(file, header=4) # Skip the first 4 lines, start from the 5th, ZERO indexed
            qbank["QDF"] = df # Value
            qbank["FILENAME"] = file # Lets store the file name as well.
            self.topics.append(qbank["TOPIC"])
            key = qbank["TOPIC"] + ";" + qbank["SUBTOPIC"]
            self.subtopics[key]  = qbank["DESCRIPTION"]
            self.subtopicsDF[key] = df
            if qbank["TOPIC"] not in self.subtopicsList.keys() :
                self.subtopicsList[qbank["TOPIC"]] = [] # Create a list
            self.subtopicsList[qbank["TOPIC"]].append(qbank["SUBTOPIC"])
                        
            self.qBanksAll[file] = qbank
        self.topics.sort()
        
    def getTopics(self) :
            return json.dumps(self.topics)
    
    def getSubTopics(self, topic) :
            return json.dumps(self.subtopicsList[topic])

    #Gets a selection of questions from a Question Bank. Default return type is a JSON
    #Level = '0' means that questions without any level distiction to be returned
    #If exam is TRUE, answer and discription is not returned.
    def getQuestions(self, topic, subtopic, level = '0', numOfQuestions = 0, exam = True , retType = 'J') :
        key = topic + ";" + subtopic
        if key not in self.subtopicsDF.keys() :
            return json.dumps("Questions NOT FOUND")
        dft1 = self.subtopicsDF[key]
        df = dft1.dropna() # Drop all rows where ANY value is NA EVERYTHING should be defined

        if (level != '0') : # A specific level was defined, get the subset
            df = df[df.DifficultyLevel == level]

        if numOfQuestions : # A number great than ZERO was specified
            if numOfQuestions <= len(df) :
                df = df.sample(numOfQuestions) # What happens if numofquestions is > num available? Not sure
            else :
                print("More questions requested in Test than available:%s:%s:%s" %(topic, subtopic, level))
                df = df.sample(len(df)) # What happens if numofquestions is > num available? Not sure
        
        if exam : # Exam is set as true
            df = df.iloc[:,1:6] # If true, remove answers
        else :
            df = df.iloc[:,1:]
            
        
        if retType == 'D' : #DataFrame was requested
            return (df) # Return the DataFrame Subset
        
        i = 0 
        jstr = ""
        while i < len(df) :
            jstr += df.iloc[i].to_json()
            i += 1
        return ("{" + jstr + "}")
    
   
    def getQuestionBanks(self) :
        jstr = ""
        for fn in self.qBanksAll.keys() : 
            qb = self.qBanksAll[fn]
            tl = [qb["TOPIC"],qb["SUBTOPIC"],qb["DESCRIPTION"] ]
            jstr += json.dumps(tl)
        return ("{" + jstr + "}")
    


# In[2]:

if __name__ == "__main__":
    appC = AppConfig("C:\\Users\\kambhs\\Desktop\\Projects\\OnlineExam\\OnlineExamConfig.xlsx")
    qb = QuestionBank(appC)


# In[20]:

if __name__ == "__main__":
    key = 'Life Insurance' + ";" + "Basics of Life Insurance"
    df = qb.subtopicsDF[key]
    level = "E"
    #Sample calls
    qb.getQuestions('Life Insurance', "Basics of Life Insurance",level="E", numOfQuestions = 20, retType="DF", exam=False)
    qb.getQuestions('Life Insurance', "Basics of Life Insurance", numOfQuestions=3, exam =False, retType='D')
    qb.getQuestions('Life Insurance', "Basics of Life Insurance", exam =True)
    qb.getSubTopics('Life Insurance')
    i =0
    while i < len(c) :
        print(c.iloc[i].Question)
        print(c.iloc[i].OptionA)
        print(c.iloc[i].OptionB)
        print(c.iloc[i].OptionC)
        print(c.iloc[i].OptionD)
        print(c.iloc[i].DifficultyLevel)
        print(c.iloc[i].Description)
        i += 1


# In[ ]:




# In[ ]:




# In[ ]:




# In[ ]:




# In[ ]:




# In[ ]:




# In[ ]:




# In[ ]:



