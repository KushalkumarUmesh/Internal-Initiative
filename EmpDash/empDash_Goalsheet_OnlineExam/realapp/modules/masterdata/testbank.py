"""
Project: OnlineExam
Domain: TestBank
Description: The TestBank handles all functionalities related to Test Definitions by picking-up questions from multiple
Question Banks and adding some additional attributes (passing marks, total num of questions, etc.)
The following need to be supplied. Responses are expected to be JSON or dict, depending on a global flag. JSON is for Angular2
UI. Ability to return the data as-is allows for alternate processing e.g. storing in DB
TODO:
a) List of Tests (Questions) - This is collected from all the Test-Definitions, check needs to made for Unique
    => This has not been fully implemented, need to see how to render a list
b)- List of Difficulty Levels(Questions) - Collated Test-Definitions --?
c) Generate Question-Set - Exam and non-Exam by calling the methods in the question-bank : Main Purpose

Author: K.Srinivas
16-Mar-2018
"""
import logging
import glob
import openpyxl
import pandas as pd
import json
from questionbank import QuestionBank
from appconfig import AppConfig


# This class reads reads Test Definition-XLS files collates data for future use
# Data is expected to be small enough to be read and stored in memory.
# Core Function is to return a set of questions as per the test-definition
# I use appConfig heavily. All the paths, values needed must be defined in Appconfig
class TestBank(object) :
    def __init__(self, appConf) :
        tmatch = appConf.attribute["TestMatch"] # Directory containing all the test-definitions
        tRow = 3 # appConf.attribute["TOPICROW"] # Row to find TOPIC in the XLS-file, will assume the remaining two
        tCol = 2 # appConf.attribute["TOPICROW"] 
        self.testsAll = dict() # All Tests stored
        self.testList = dict() # List of all tests, for UI purpose only
        self.qb = dict() # Question Bank
        
        for file in glob.glob(tmatch):  # For each of the question banks
            tbank = dict() # Define a question Bank
            #Get title, Get Subtitle, Get Desc
            conf_wb = openpyxl.load_workbook(file)
            ws = conf_wb.worksheets[0] # Assume only 1 worksheet
            if len(conf_wb.worksheets) > 1 : # Error Check
                print("More than one worksheets found in %s" % (file))
                
            c_row =  tRow
            # Test Name  = Description
            tn = ws.cell(row=c_row, column=tCol).value # Name of the Test
            if tn in self.testsAll.keys() :
                print("Duplicate Test Name found in: %s" % (file))
                exit(1) # Not sure if this works
            td = ws.cell(row=c_row, column=(tCol+1)).value # Test Description
            c_row +=1  
            tnumQ = ws.cell(row=c_row, column=tCol).value # Number of questions
            c_row +=1 
            tdiff = ws.cell(row=c_row, column=tCol).value # Difficulty level
            c_row +=1
            tpass = ws.cell(row=c_row, column=tCol).value # Pass-number
            c_row +=1
            tset1_topic =  ws.cell(row=c_row, column=tCol).value # Topic
            tset1_subtopic =  ws.cell(row=c_row, column=tCol+1).value # Subtopic
            tset1_numq = ws.cell(row=c_row, column=tCol+2).value # Num of questions from this set
            tset1_diff =  ws.cell(row=c_row, column=tCol+3).value # diff-level from this set
            
            tname = dict()
            tname["TestName"] = tn
            tname["DESCRIPTION"] = td
            tname["NoOfQuestions"] = tnumQ
            tname["DifficultyLevel"] = tdiff
            tname["PassNum"] = tpass
            tname["Qbank1"] = [tset1_topic,tset1_subtopic, tset1_numq, tset1_diff ]
            tname["FileName"] = file # Name of the File we read this from, for debugging/error reporting purposes

            self.testsAll[tn] = tname # Our final collection
            self.testList[tn] = td  # Create a list of Test-Names, with description as the value
                         

    # This is needed so as to 
    def setQuestionBank(self, qb) :
        self.qb = qb

    def getTestDesc(self, testName) :
        tData = self.testsAll[testName] # Get the Dict of the Test
        return tData["DESCRIPTION"]

    def getTestNoOfQuestions(self, testName) :
        tData = self.testsAll[testName] # Get the Dict of the Test
        return tData["NoOfQuestions"]
    
    def getTestPassNum(self, testName) :
        tData = self.testsAll[testName] # Get the Dict of the Test
        return tData["PassNum"]

    def getTestDifficultyLevel(self, testName) :
        tData = self.testsAll[testName] # Get the Dict of the Test
        return tData["DifficultyLevel"]
        
    def getTestList(self, retType = 'D') :
        if retType == 'D' :
            return(self.testList)
        rt = dict()
        for key in self.testList.keys() :
            rt[key] =  key
        return (json.dumps(rt))

    """
    # Misleading - this gets the Question-Bank specific data, that we don't need
    def getTestData(self, testName, retType = 'J') :
        tData = self.testsAll[testName] # Get the Dict of the Test
        return ([tData["Qbank1"][0], tData["Qbank1"][1], tData["Qbank1"][2], tData["Qbank1"][3]])

        if retType == 'D' :
            return(self.testList)
        rt = dict()
        for key in self.testList.keys() :
            rt[key] =  key
        return (json.dumps(rt))
    """
        
    def getQuestions(self, testName,retType = 'D',  exam = False) :
        tData = self.testsAll[testName] # Get the Dict of the Test
        #tData["Qbank1"][0], tData["Qbank1"][0] 
        #return (tData)
        return (self.qb.getQuestions(tData["Qbank1"][0], tData["Qbank1"][1], level=tData["Qbank1"][3], \
                 numOfQuestions = tData["Qbank1"][2], retType=retType, exam=exam))

    def checkConsistancy(self) :
        errFound = False
        print("Question Bank: Consistancy Check... ")
        if not self.qb :
            print("Consistancy Check called without setting Question Bank")
        for  t in self.testsAll.keys() : # Scan all tests
            tData =  self.testsAll[t]
            qs = self.qb.getQuestions(tData["Qbank1"][0], tData["Qbank1"][1], level=tData["Qbank1"][3], \
                 numOfQuestions = tData["Qbank1"][2], retType='D', exam=True)
            if len(qs) != tData["NoOfQuestions"] :
                errFound = True   
                print (tData["TestName"] + " Does not have enough questions or Questio Bank specified does not exist." )
        if errFound :
            exit(1)
        else :
            return True

if __name__ == "__main__":
    appC = AppConfig("C:\\Users\\kambhs\\Desktop\\Projects\\OnlineExam\\OnlineExamConfig.xlsx")
    qb = QuestionBank(appC)
    testBank = TestBank(appC)
    testBank.setQuestionBank(qb)
    testBank.getQuestions('Java Level 1', exam=True)
    testBank.getTestList(retType = 'D')
